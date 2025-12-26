import os
import sqlite3
import calendar
from datetime import datetime, time, timedelta, date

from openpyxl import Workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes

# -----------------------------
# CONFIG
# -----------------------------
# Opzione A (consigliata): variabile ambiente TOKEN
TOKEN = os.getenv("TOKEN")

if not TOKEN:
    raise ValueError("ERRORE: TOKEN non impostato. Imposta la variabile ambiente TOKEN o scrivilo nel file.")

ADMIN_CHAT_ID = 6970805908  # <-- il TUO user id Telegram (admin)

DB_PATH = "presenze.db"
EXCEL_ROLLING_FILE = "resoconto_ultimi_2_mesi.xlsx"


# -----------------------------
# DATABASE
# -----------------------------
conn = sqlite3.connect(DB_PATH, check_same_thread=False)
conn.execute("""
CREATE TABLE IF NOT EXISTS presenze (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    nome TEXT,
    azione TEXT,
    timestamp TEXT,
    lat REAL,
    lon REAL
)
""")
conn.commit()


def registra_presenza(user_id: int, nome: str, azione: str, lat=None, lon=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        "INSERT INTO presenze (user_id, nome, azione, timestamp, lat, lon) VALUES (?, ?, ?, ?, ?, ?)",
        (user_id, nome, azione, timestamp, lat, lon)
    )
    conn.commit()


# -----------------------------
# UTIL
# -----------------------------
def is_admin(update: Update) -> bool:
    return bool(update.effective_user and update.effective_user.id == ADMIN_CHAT_ID)


def _format_ore_minuti(seconds: float):
    ore = int(seconds // 3600)
    minuti = int((seconds % 3600) // 60)
    return ore, minuti


def start_of_week(dt: datetime) -> datetime:
    start = dt - timedelta(days=dt.weekday())  # lunedì
    return start.replace(hour=0, minute=0, second=0, microsecond=0)


# -----------------------------
# PAIRING ENTRATA -> USCITA (ROBUSTO)
# -----------------------------
def pair_sessions(records):
    """
    Crea coppie (ENTRATA, USCITA) valide.

    Regole:
    - Se manca USCITA: NON conta nulla (sessione ignorata).
    - USCITA valida solo se nello STESSO giorno della ENTRATA.
    - Se arriva un'altra ENTRATA mentre sei già "dentro", si scarta l'entrata precedente (non chiusa).
    - USCITA senza ENTRATA: ignorata.
    """
    sessions = []
    current_in = None

    for azione, ts in records:
        ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")

        if azione == "ENTRATA":
            current_in = ts  # tengo l'ultima entrata (scarta eventuale entrata precedente non chiusa)

        elif azione == "USCITA":
            if current_in is None:
                continue  # uscita senza entrata

            if ts.date() != current_in.date():
                # NON accoppiare tra giorni diversi
                current_in = None
                continue

            if ts > current_in:
                sessions.append((current_in, ts))
            current_in = None

    return sessions


# -----------------------------
# CALCOLO ORE NORMALI + STRAORDINARI
# -----------------------------
def calcola_straordinari(records):
    sessions = pair_sessions(records)
    totale_normali = 0
    totale_extra = 0

    for inizio, fine in sessions:
        giorno = inizio.weekday()  # 0 lun ... 5 sab

        if giorno <= 4:  # lun-ven
            inizio_std = datetime.combine(inizio.date(), time(7, 0))
            fine_std = datetime.combine(inizio.date(), time(16, 30))
        else:  # sabato
            inizio_std = datetime.combine(inizio.date(), time(7, 30))
            fine_std = datetime.combine(inizio.date(), time(13, 0))

        # Normali
        normale_inizio = max(inizio, inizio_std)
        normale_fine = min(fine, fine_std)
        if normale_fine > normale_inizio:
            totale_normali += (normale_fine - normale_inizio).total_seconds()

        # Straordinari (prima + dopo)
        extra_prima = max(0, (inizio_std - inizio).total_seconds())
        extra_dopo = max(0, (fine - fine_std).total_seconds())
        totale_extra += (extra_prima + extra_dopo)

    return totale_normali, totale_extra


def calcola_extra_per_giorno(records):
    sessions = pair_sessions(records)
    extra_by_date = {}

    for inizio, fine in sessions:
        giorno = inizio.weekday()

        if giorno <= 4:
            inizio_std = datetime.combine(inizio.date(), time(7, 0))
            fine_std = datetime.combine(inizio.date(), time(16, 30))
        else:
            inizio_std = datetime.combine(inizio.date(), time(7, 30))
            fine_std = datetime.combine(inizio.date(), time(13, 0))

        extra_prima = max(0, (inizio_std - inizio).total_seconds())
        extra_dopo = max(0, (fine - fine_std).total_seconds())
        extra = extra_prima + extra_dopo

        d = inizio.date()
        extra_by_date[d] = extra_by_date.get(d, 0) + extra

    return extra_by_date


# -----------------------------
# COMANDI BASE
# -----------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Ciao! Sono il bot presenze.\n"
        "Comandi:\n"
        "• /entra – registra entrata\n"
        "• /esci – registra uscita\n"
        "• /ore – ore lavorate (normali + straordinari)\n"
        "• /straordinari – ore extra\n"
        "• /export – export presenze (tutto)\n"
        "\nAdmin:\n"
        "• /settimana – dettaglio settimana corrente\n"
        "• /settimane [N] – totali divisi per settimana\n"
        "• /excel2mesi – Excel ultimi 2 mesi\n"
    )


async def entra(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    registra_presenza(user.id, user.full_name, "ENTRATA")
    await update.message.reply_text("Entrata registrata ✔")


async def esci(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    registra_presenza(user.id, user.full_name, "USCITA")
    await update.message.reply_text("Uscita registrata ✔")


async def straordinari(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    records = conn.execute(
        "SELECT azione, timestamp FROM presenze WHERE user_id = ? ORDER BY timestamp ASC",
        (user.id,)
    ).fetchall()

    if not records:
        await update.message.reply_text("Nessuna presenza registrata.")
        return

    _, extra = calcola_straordinari(records)
    h, m = _format_ore_minuti(extra)
    await update.message.reply_text(f"🔥 Straordinari totali: {h} ore e {m} minuti")


async def ore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    records = conn.execute(
        "SELECT azione, timestamp FROM presenze WHERE user_id = ? ORDER BY timestamp ASC",
        (user.id,)
    ).fetchall()

    if not records:
        await update.message.reply_text("Nessuna presenza registrata.")
        return

    norm, extra = calcola_straordinari(records)
    tot = norm + extra

    hn, mn = _format_ore_minuti(norm)
    he, me = _format_ore_minuti(extra)
    ht, mt = _format_ore_minuti(tot)

    await update.message.reply_text(
        f"🕒 Ore normali: {hn}h {mn}m\n"
        f"🔥 Straordinari: {he}h {me}m\n"
        f"✅ Totale: {ht}h {mt}m"
    )


# -----------------------------
# ADMIN: SETTIMANA CORRENTE (DETTAGLIO)
# -----------------------------
async def settimana(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("⛔ Comando riservato all’amministratore.")
        return

    now = datetime.now()
    start_w = start_of_week(now)
    end = now + timedelta(seconds=1)

    utenti = conn.execute(
        "SELECT DISTINCT user_id, nome FROM presenze ORDER BY nome COLLATE NOCASE"
    ).fetchall()

    days = []
    d = start_w.date()
    while d <= now.date():
        days.append(d)
        d = (datetime.combine(d, time(0, 0)) + timedelta(days=1)).date()

    text = f"📋 *Straordinari settimana* (da {start_w.strftime('%d/%m')} a {now.strftime('%d/%m')})\n\n"

    for user_id, nome in utenti:
        rec = conn.execute(
            "SELECT azione, timestamp FROM presenze "
            "WHERE user_id = ? AND timestamp >= ? AND timestamp < ? "
            "ORDER BY timestamp",
            (user_id, start_w.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S"))
        ).fetchall()

        extra_by_day = calcola_extra_per_giorno(rec)

        total = 0
        righe = []
        for day in days:
            sec = extra_by_day.get(day, 0)
            total += sec
            h, m = _format_ore_minuti(sec)
            righe.append(f"  • {day.strftime('%a %d/%m')}: {h}h {m}m")

        ht, mt = _format_ore_minuti(total)
        text += f"*{nome}* — Totale: *{ht}h {mt}m*\n" + "\n".join(righe) + "\n\n"

    await update.message.reply_text(text, parse_mode="Markdown")


# -----------------------------
# ADMIN: PIÙ SETTIMANE (DIVISE) - TOTALI PER SETTIMANA
# -----------------------------
async def settimane(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("⛔ Comando riservato all’amministratore.")
        return

    n = 4
    if context.args:
        try:
            n = max(1, min(12, int(context.args[0])))
        except Exception:
            n = 4

    now = datetime.now()
    this_week_start = start_of_week(now)

    utenti = conn.execute(
        "SELECT DISTINCT user_id, nome FROM presenze ORDER BY nome COLLATE NOCASE"
    ).fetchall()

    text = f"📚 *Straordinari ultime {n} settimane* (divise)\n\n"

    for w in range(n):
        week_start = this_week_start - timedelta(days=7 * w)
        week_end = week_start + timedelta(days=7)

        # giorni lun->dom (solo per sommare)
        days = []
        d = week_start.date()
        while d < week_end.date():
            days.append(d)
            d = (datetime.combine(d, time(0, 0)) + timedelta(days=1)).date()

        text += f"🗓 *Settimana {week_start.strftime('%d/%m')}–{(week_end - timedelta(days=1)).strftime('%d/%m')}*\n"

        for user_id, nome in utenti:
            rec = conn.execute(
                "SELECT azione, timestamp FROM presenze "
                "WHERE user_id = ? AND timestamp >= ? AND timestamp < ? "
                "ORDER BY timestamp",
                (user_id, week_start.strftime("%Y-%m-%d %H:%M:%S"), week_end.strftime("%Y-%m-%d %H:%M:%S"))
            ).fetchall()

            extra_by_day = calcola_extra_per_giorno(rec)
            total = sum(extra_by_day.get(day, 0) for day in days)
            ht, mt = _format_ore_minuti(total)

            text += f"• {nome}: *{ht}h {mt}m*\n"

        text += "\n"

    await update.message.reply_text(text, parse_mode="Markdown")


# -----------------------------
# EXPORT EXCEL (TUTTO)
# -----------------------------
async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Presenze"
    ws.append(["ID", "User ID", "Nome", "Azione", "Timestamp", "Lat", "Lon"])

    rows = conn.execute("SELECT * FROM presenze ORDER BY id ASC").fetchall()
    for row in rows:
        ws.append(row)

    filename = "presenze.xlsx"
    wb.save(filename)

    await update.message.reply_document(open(filename, "rb"))


# -----------------------------
# EXCEL ULTIMI 2 MESI (ADMIN)
# -----------------------------
def add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    last = calendar.monthrange(y, m)[1]
    day = min(d.day, last)
    return date(y, m, day)


def two_month_range(now: datetime):
    start_month = add_months(now.date().replace(day=1), -1)
    end_month = add_months(now.date().replace(day=1), 1)  # primo giorno mese prossimo
    start_dt = datetime.combine(start_month, time(0, 0))
    end_dt = datetime.combine(end_month, time(0, 0))
    return start_dt, end_dt


def generate_excel_last_2_months():
    now = datetime.now()
    start_dt, end_dt = two_month_range(now)

    rows = conn.execute(
        "SELECT user_id, nome, azione, timestamp FROM presenze "
        "WHERE timestamp >= ? AND timestamp < ? "
        "ORDER BY nome COLLATE NOCASE, timestamp ASC",
        (start_dt.strftime("%Y-%m-%d %H:%M:%S"), end_dt.strftime("%Y-%m-%d %H:%M:%S"))
    ).fetchall()

    by_user = {}
    for user_id, nome, azione, ts in rows:
        by_user.setdefault((user_id, nome), []).append((azione, ts))

    days = []
    d = start_dt.date()
    while d < end_dt.date():
        days.append(d)
        d = (datetime.combine(d, time(0, 0)) + timedelta(days=1)).date()

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"
    ws.append(["Dipendente"] + [d.strftime("%d/%m") for d in days] + ["Totale (h)"])

    for (user_id, nome), records in sorted(by_user.items(), key=lambda x: x[0][1].lower()):
        extra_by_day = calcola_extra_per_giorno(records)

        total_sec = 0
        row = [nome]
        for d in days:
            sec = extra_by_day.get(d, 0)
            total_sec += sec
            h, m = _format_ore_minuti(sec)
            row.append("" if sec == 0 else f"{h:02d}:{m:02d}")

        ht, mt = _format_ore_minuti(total_sec)
        row.append(f"{ht:02d}:{mt:02d}")
        ws.append(row)

    ws2 = wb.create_sheet("Log")
    ws2.append(["User ID", "Nome", "Azione", "Timestamp"])
    for user_id, nome, azione, ts in rows:
        ws2.append([user_id, nome, azione, ts])

    wb.save(EXCEL_ROLLING_FILE)


async def excel2mesi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("⛔ Comando riservato all’amministratore.")
        return

    generate_excel_last_2_months()
    await update.message.reply_document(open(EXCEL_ROLLING_FILE, "rb"))


# -----------------------------
# REPORT AUTOMATICO SETTIMANALE (settimana corrente)
# -----------------------------
async def report_settimanale(app):
    now = datetime.now()
    start_w = start_of_week(now)
    end_w = start_w + timedelta(days=7)

    utenti = conn.execute(
        "SELECT DISTINCT user_id, nome FROM presenze ORDER BY nome COLLATE NOCASE"
    ).fetchall()

    days = []
    d = start_w.date()
    while d < end_w.date():
        days.append(d)
        d = (datetime.combine(d, time(0, 0)) + timedelta(days=1)).date()

    text = f"📊 *Report straordinari settimana* ({start_w.strftime('%d/%m')}–{(end_w - timedelta(days=1)).strftime('%d/%m')})\n\n"

    for user_id, nome in utenti:
        rec = conn.execute(
            "SELECT azione, timestamp FROM presenze "
            "WHERE user_id = ? AND timestamp >= ? AND timestamp < ? "
            "ORDER BY timestamp",
            (user_id, start_w.strftime("%Y-%m-%d %H:%M:%S"), end_w.strftime("%Y-%m-%d %H:%M:%S"))
        ).fetchall()

        extra_by_day = calcola_extra_per_giorno(rec)

        total = 0
        righe = []
        for day in days:
            sec = extra_by_day.get(day, 0)
            total += sec
            h, m = _format_ore_minuti(sec)
            righe.append(f"  • {day.strftime('%a %d/%m')}: {h}h {m}m")

        ht, mt = _format_ore_minuti(total)
        text += f"*{nome}* — Totale: *{ht}h {mt}m*\n" + "\n".join(righe) + "\n\n"

    await app.bot.send_message(ADMIN_CHAT_ID, text, parse_mode="Markdown")


async def start_scheduler(app):
    scheduler = AsyncIOScheduler()
    # sabato alle 13:05
    scheduler.add_job(
        report_settimanale, "cron",
        day_of_week="sat", hour=13, minute=5,
        args=[app]
    )
    scheduler.start()


# -----------------------------
# MAIN
# -----------------------------
def main():
    app = ApplicationBuilder().token(TOKEN).post_init(start_scheduler).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("entra", entra))
    app.add_handler(CommandHandler("esci", esci))
    app.add_handler(CommandHandler("ore", ore))
    app.add_handler(CommandHandler("straordinari", straordinari))

    app.add_handler(CommandHandler("export", export))
    app.add_handler(CommandHandler("settimana", settimana))
    app.add_handler(CommandHandler("settimane", settimane))
    app.add_handler(CommandHandler("excel2mesi", excel2mesi))

    print("Bot avviato…")
    app.run_polling()


if __name__ == "__main__":
    main()



